import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import xml.etree.ElementTree as ET
import openpyxl

class ProfessorReg:
    def __init__(self, root):
        self.root = root       
        self.root.title("Administrador de Cursos")
        self.root.geometry("700x400")
        self.root['bg'] = '#D8B175'
        self.root.iconbitmap('Usac_logo.ico')

        # Crear una lista para almacenar los nombres de los catedráticos
        self.catedraticos = []

        # Cargar el archivo XML
        self.tree = ET.parse('cursos.xml')
        self.root_cursos = self.tree.getroot()

        # Iterar sobre los elementos XML y recopilar los nombres de los catedráticos únicos
        for curso in self.root_cursos:            
            catedratico = curso.find('Catedratico').text
            print(catedratico)
            if catedratico not in self.catedraticos:
                self.catedraticos.append(catedratico)

        print(self.catedraticos)

        # Crear la lista desplegable con los nombres de los catedráticos
        self.professor_label = tk.Label(self.root, text="Catedrático:")
        self.professor_label.grid(row=0, column=2, columnspan=2, padx=10, pady=10)
        self.professor_label['bg'] = '#D8B175'
        self.professor = ttk.Combobox(self.root, values=self.catedraticos)
        self.professor.grid(row=0, column=4, padx=10, pady=10)

        self.tree = ET.parse('cursos.xml')
        self.root_cursos = self.tree.getroot()

        self.cursos = []

        self.cursos_label = tk.Label(self.root, text="Cursos:")
        self.cursos_label.grid(row=1, column=1, padx=10, pady=10)
        self.cursos_listbox = tk.Listbox(self.root)
        self.cursos_listbox.grid(row=1, column=2, padx=10, pady=10)

        self.horario_label = tk.Label(self.root, text="Horarios:")
        self.horario_label.grid(row=1, column=3, padx=10, pady=10)
        self.horarios_listbox = tk.Listbox(self.root)
        self.horarios_listbox.grid(row=1, column=4, padx=10, pady=10)

        self.students_label = tk.Label(self.root, text="Estudiantes:")
        self.students_label.grid(row=1, column=5, padx=10, pady=10)
        self.students_listbox = tk.Listbox(self.root)
        self.students_listbox.grid(row=1, column=6, padx=10, pady=10)

        self.register_button = tk.Button(self.root, text="Cargar cursos", command=self.view_courses)
        self.register_button.grid(row=6, column=2, padx=10, pady=10)

        self.salir_button = tk.Button(self.root, text="Salir", command=self.logout)
        self.salir_button.grid(row=6, column=4, padx=10, pady=10)

        self.estudiantes_button = tk.Button(self.root, text="Ver estudiantes", command=self.view_students)
        self.estudiantes_button.grid(row=6, column=6, padx=10, pady=10)  
        

    def show_message(self, message):
        messagebox.showinfo("Éxito", message)

    def view_courses(self):
        catedratico = self.professor.get()

        if not catedratico:
            self.show_message("Selecciona un catedrático")
            return

        # Limpia las listas de cursos y horarios antes de cargar los nuevos datos
        self.cursos_listbox.delete(0, tk.END)
        self.horarios_listbox.delete(0, tk.END)

        # Itera sobre los elementos XML y muestra los cursos y horarios del catedrático seleccionado
        for curso in self.root_cursos:
            if curso.find('Catedratico').text == catedratico:
                nombre_curso = curso.find('Nombre').text
                horario_curso = curso.find('Horario').text
                self.cursos_listbox.insert(tk.END, nombre_curso)
                self.horarios_listbox.insert(tk.END, horario_curso)

    def view_students(self):
        catedratico = self.professor.get()

        if not catedratico:
            self.show_message("Selecciona un catedrático")
            return

        # Lee el archivo XML de cursos
        cursos_tree = ET.parse('cursos.xml')
        cursos_root = cursos_tree.getroot()

        # Obtén una lista de cursos impartidos por el catedrático seleccionado
        cursos_catedratico = []

        for curso in cursos_root:
            if curso.find('Catedratico').text == catedratico:
                nombre_curso = curso.find('Nombre').text
                cursos_catedratico.append(nombre_curso)        

        if not cursos_catedratico:
            self.show_message("El catedrático seleccionado no imparte cursos.")
            return
        
        # Borra los elementos anteriores en el Listbox de estudiantes
        self.students_listbox.delete(0, tk.END)
        
        # Lee el archivo XML de estudiantes
        students_tree = ET.parse('students.xml')
        students_root = students_tree.getroot()

        # Lista para almacenar a los estudiantes inscritos en los cursos del catedrático
        estudiantes_inscritos = []

        # Itera sobre los estudiantes y sus cursos
        for student in students_root:
            cursos_estudiante = student.find('Cursos')              

            if cursos_estudiante is not None:
                for curso_element in cursos_estudiante:
                    if curso_element.text in cursos_catedratico:
                        nombre_estudiante = student.find('Nombre').text
                        apellidos_estudiante = student.find('Apellido').text
                        estudiantes_inscritos.append(f"{nombre_estudiante} {apellidos_estudiante}")

        if not estudiantes_inscritos:
            self.show_message("No hay estudiantes inscritos en las materias del catedrático seleccionado.")
        else:
            # alumnos_str = "\n".join(estudiantes_inscritos)
            # messagebox.showinfo("Estudiantes Inscritos", alumnos_str)
            # Inserta los nombres de los estudiantes en el Listbox de estudiantes
            for estudiante in estudiantes_inscritos:
                self.students_listbox.insert(tk.END, estudiante)
            
            # Guarda los datos en un archivo Excel
            self.save_to_excel(catedratico, cursos_catedratico, estudiantes_inscritos)

    def save_to_excel(self, catedratico, cursos, estudiantes):
        # Crea un nuevo archivo de Excel o abre un archivo existente
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Datos del Catedrático"

        # Escribe los encabezados
        sheet['A1'] = 'Catedrático'
        sheet['B1'] = 'Curso'
        sheet['C1'] = 'Horario'
        sheet['D1'] = 'Estudiantes'

        # Escribe los datos
        for i, curso in enumerate(cursos):
            sheet.cell(row=i + 2, column=1, value=catedratico)
            sheet.cell(row=i + 2, column=2, value=curso)
            sheet.cell(row=i + 2, column=3, value="Horario correspondiente")  # Reemplaza con el horario real
            sheet.cell(row=i + 2, column=4, value=estudiantes[i])

        # Guarda el archivo de Excel
        workbook.save('datos_catedratico.xlsx')
        self.show_message("Los datos se han guardado en 'datos_catedratico.xlsx'")


    def logout(self):
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    professor_panel = ProfessorReg(root)
    root.mainloop()
